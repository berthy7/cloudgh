from ...operaciones.bitacora.managers import *
from ...personal.persona.models import *
from ..rol.models import *

from server.common.managers import SuperManager, Error
from .models import *
from sqlalchemy.sql import func
from random import *

import string
import random
import hashlib

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side

class UsuarioManager(SuperManager):

    def __init__(self, db):
        super().__init__(Usuario, db)

    def usuario_excel(self, ):
        cname = "Usuarios.xlsx"

        usuarios = self.db.query(self.entity).filter(self.entity.username != "admin").order_by(self.entity.id.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'a'

        indice = 1

        ws['A' + str(indice)] = 'ID'
        ws['B' + str(indice)] = 'USERNAME'
        ws['C' + str(indice)] = 'PASSWORD'
        ws['D' + str(indice)] = 'TOKEN'
        ws['E' + str(indice)] = 'AUTENTICACION'
        ws['F' + str(indice)] = 'ROL'
        ws['G' + str(indice)] = 'PERSONAL'
        ws['H' + str(indice)] = 'ESTADO'


        for i in usuarios:
            indice = indice + 1

            codigo = ""

            if i.fkpersona:
                codigo = i.persona.empleado[0].codigo

            ws['A' + str(indice)] = i.id
            ws['B' + str(indice)] = i.username
            ws['C' + str(indice)] = i.password
            ws['D' + str(indice)] = i.token
            ws['E' + str(indice)] = i.autenticacion
            ws['F' + str(indice)] = i.rol.nombre
            ws['G' + str(indice)] = codigo
            ws['H' + str(indice)] = i.enabled

        wb.save("server/common/resources/downloads/" + cname)
        return cname

    def importar_excel(self, cname, user, ip):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['ID', 'USERNAME', 'PASSWORD', 'TOKEN', 'AUTENTICACION', 'ROL', 'PERSONAL', 'ESTADO']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                for row in ws.iter_rows(min_row=2):

                    if row[indices['USERNAME']].value is not None:

                        query = self.db.query(Usuario).filter(Usuario.username == row[indices['USERNAME']].value).first()

                        if not query:
                            fkpersona = None
                            if row[indices['PERSONAL']].value != "":
                                query_persona = self.db.query(Empleado).filter(Empleado.codigo == row[indices['PERSONAL']].value).first()
                                if query_persona:
                                    fkpersona = query_persona.fkpersona

                            query_rol = self.db.query(Rol).filter(Rol.nombre == row[indices['ROL']].value).first()

                            user = Usuario(username=row[indices['USERNAME']].value, password=row[indices['PASSWORD']].value, token=row[indices['TOKEN']].value,
                                            autenticacion=row[indices['AUTENTICACION']].value, fkrol=query_rol.id, fkpersona=fkpersona,
                                            enabled=row[indices['ESTADO']].value)

                            self.db.merge(user)
                            self.db.flush()

                    else:

                        self.db.rollback()
                        return {'message': 'Hay Columnas vacias', 'success': False}

                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except Exception as e:
            self.db.rollback()
            if 'UNIQUE constraint' in str(e):
                return {'message': 'duplicado', 'success': False}
            if 'UNIQUE constraint failed' in str(e):
                return {'message': 'codigo duplicado', 'success': False}
            return {'message': str(e), 'success': False}

    def obtener_administrador(self):
        return self.db.query(Usuario).filter(Usuario.nombre == "Administrador").one()

    def name_role(self, rol):
        role = self.db.query(Rol).filter_by(id=rol).first()
        nombre_rol = role.name
        return nombre_rol

    def get_random_string(self):
        random_list = []
        for i in range(8):
            random_list.append(random.choice(string.ascii_uppercase + string.digits))
        return ''.join(random_list)

    def insert(self, usuario):
        user = self.db.query(Usuario).filter(Usuario.username == usuario.username).first()

        if user:

            if user.fkpersona:

                return dict(respuesta=False, Mensaje="Ya se le creo un usuario al personal "+ user.persona.fullname)
            else:
                return dict(respuesta=False, Mensaje="Ingrese otro Nombre de usuario")

        else:
            usuario.password = hashlib.sha512(usuario.password.encode()).hexdigest()
            codigo = self.get_random_string()
            Usuario.codigo = codigo
            fecha = BitacoraManager(self.db).fecha_actual()
            b = Bitacora(fkusuario=usuario.user_id, ip=usuario.ip, accion="Se registró un usuario.", fecha=fecha)
            super().insert(b)
            u = super().insert(usuario)
            return dict(respuesta=True, Mensaje="Insertado Correctamente")
        
    def registrar_usuarios(self, usuarios):
        
        for usuario in usuarios['personas_arbol']:
            user = self.db.query(Usuario).filter(Usuario.fkpersona == usuario['id_persona']).first()

            if user:
                print("La persona "+user.persona.fullname+" ya tiene usuario")
               
            else:
                persona = self.db.query(Persona).filter(Persona.id == usuario['id_persona']).first()
                password = "personal2021"

                if persona.empleado[0].email:
                    correo = persona.empleado[0].email
                else:
                    correo = persona.nombres
                
                diccionary_user = Usuario(username=correo, password= password, fkrol=usuarios['fkrol'], fkpersona=persona.id)

                diccionary_user.password = hashlib.sha512(diccionary_user.password.encode()).hexdigest()
                fecha = BitacoraManager(self.db).fecha_actual()
                b = Bitacora(fkusuario=usuarios['user'], ip=usuarios['ip'], accion="Se creo el usuario: "+ diccionary_user.username, fecha=fecha)
                super().insert(b)
                super().insert(diccionary_user)


    def update(self, usuario):

        if not usuario.password or usuario.password == '':
            usuario.password = (self.db.query(Usuario.password)
                .filter(Usuario.id == usuario.id).first())[0]
        else:
            usuario.password = hashlib.sha512(usuario.password.encode()).hexdigest()

        fecha = BitacoraManager(self.db).fecha_actual()
        a = super().update(usuario)
        b = Bitacora(fkusuario=usuario.user_id, ip=usuario.ip, accion="Modificó Usuario.", fecha=fecha, tabla="cb_usuarios_usuario", identificador=a.id)
        super().insert(b)

        return a

    def update_users(self, emailprev, emailnew, nameprev, namenew):
        u = self.db.query(Usuario).filter(Usuario.correo == emailprev).one()

        if u:
            ap_user = u.apellido
            result = nameprev.index(ap_user)
            print(result)
            u.correo = emailnew

    def delete_user(self, id, enable, Usuariocr, ip):
        x = self.db.query(Usuario).filter(Usuario.id == id).one()

        if enable == True:
            r = self.db.query(Rol).filter(Rol.id == x.fkrol).one()
            if r.enabled:
                x.enabled = enable
            else:
                return False
            message = "Se habilitó un usuario."
        else:
            x.enabled = enable
            message = "Se inhabilitó un usuario."

        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=Usuariocr, ip=ip, accion=message, fecha=fecha)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()
        return True

    def activate_Usuarios(self, id, Usuario, ip):
        x = self.db.query(Usuario).filter(Usuario.id == id).one()
        x.enabled = 1
        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=Usuario, ip=ip, accion="Se activó un usuario.", fecha=fecha)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()

    def get_privileges(self, id, route):
        parent_module = self.db.query(Modulo).join(Rol.modulos).join(Usuario).            \
            filter(Modulo.route == route).\
            filter(Usuario.id == id).\
            filter(Usuario.enabled).\
            first()
        if not parent_module:
            return dict()
        modules = self.db.query(Modulo).\
            join(Rol.modulos).join(Usuario).\
            filter(Modulo.fkmodulo == parent_module.id).\
            filter(Usuario.id == id).\
            filter(Usuario.enabled)
        privileges = {parent_module.name: parent_module}
        for module in modules:
            privileges[module.name] = module
        return privileges

    def list_all(self):
        return dict(objects=self.db.query(Usuario).filter(Usuario.fkrol == Rol.id).filter(Rol.nombre != "SUPER ADMINISTRADOR").distinct())

    def has_access(self, id, route):
        aux = self.db.query(Usuario.id).\
            join(Rol).join(Acceso).join(Modulo).\
            filter(Usuario.id == id).\
            filter(Modulo.route == route).\
            filter(Usuario.enabled).\
            all()
        return len(aux) != 0

    def get_page(self, page_nr=1, max_entries=10, like_search=None, order_by=None, ascendant=True, query=None):
        query = self.db.query(Usuario).join(Rol).filter(Rol.id > 1)
        return super().get_page(page_nr, max_entries, like_search, order_by, ascendant, query)

    def login_Usuario(self, username, password):
        password = hashlib.sha512(password.encode()).hexdigest()
        return self.db.query(Usuario).filter(Usuario.username == username).filter(Usuario.password == password).filter(
            Usuario.enabled == 1)

    def get_userById(self, id):
        return dict(profile=self.db.query(Usuario).filter(Usuario.id == id).first())

    def obtener_diccionario_usuario(self, id):
        usuario = self.db.query(Usuario).filter(Usuario.id == id).first()

        if usuario.fkpersona:
            nombre = usuario.persona.fullname
            correo = usuario.persona.empleado[0].email
            autenticacion = usuario.autenticacion
        else:
            nombre = "---------"
            correo = "---------"
            autenticacion = usuario.autenticacion

        return dict(id=usuario.id, username=usuario.username,rol=usuario.rol.nombre,nombre = nombre,correo=correo,autenticacion=autenticacion)

    def update_password(self, Usuario):
        Usuario.password = hashlib.sha512(Usuario.password.encode()).hexdigest()
        return super().update(Usuario)

    def modificar_contraseña(self, id, new_pass, idUsuario,ip):
        user = self.db.query(Usuario).filter(Usuario.id == id).first()

        if user:
            user.password = hashlib.sha512(new_pass.encode()).hexdigest()


            fecha = BitacoraManager(self.db).fecha_actual()
            b = Bitacora(fkusuario=idUsuario, ip=ip, accion="Se modifico contraseña de usuario.", fecha=fecha)
            super().insert(b)
            u = super().insert(user)
            return dict(respuesta=True, Mensaje="Modificado Correctamente")

    def get_by_password(self, Usuario_id, password):
        return self.db.query(Usuario).filter(Usuario.id == Usuario_id). \
            filter(Usuario.password == hashlib.sha512(str(password).encode()).hexdigest()).first()

    def get_by_pass(self, Usuario_id):
        return self.db.query(Usuario).filter(Usuario.id == Usuario_id).first()

    def update_profile(self, Usuarioprf, ip):
        usuario = self.db.query(Usuario).filter_by(id=Usuarioprf.id).first()
        usuario.username = Usuarioprf.username
        self.db.merge(usuario)
        b = Bitacora(fkusuario=usuario.id, ip=ip, accion="Se actualizó perfil de usuario.", fecha=fecha_zona, tabla='cb_usuarios_usuario', identificador=usuario.id)
        super().insert(b)
        self.db.commit()
        return usuario

    def validar_usuario(self, username, password):
        password = hashlib.sha512(password.encode()).hexdigest()
        return self.db.query(func.count(Usuario.id)).filter(Usuario.username == username).filter(
            Usuario.enabled == True).filter(Usuario.password == password).scalar()

    def validar_usuario_sesion(self, codigo, usuario):
        return self.db.query(func.count(Usuario.id)).filter(Usuario.codigo == codigo).filter(
            Usuario.enabled == True).filter(Usuario.id == usuario).scalar()

    def activate_Usuario(self, usuario):
        usuario = self.db.query(Usuario).filter_by(id=usuario).first()
        usuario.activo = 1
        self.db.merge(usuario)
        self.db.commit()

    def update_codigo(self, usuario):
        x = self.db.query(Usuario).filter(Usuario.id == usuario).one()
        x.activo = 0
        x.codigo = self.get_random_string()
        x.token = "Sin Token"
        self.db.commit()
        self.db.close()
        return x

    def listar_todo(self, id):
        return self.db.query(Usuario).filter(Usuario.enabled == True).filter(Usuario.id == id)

    def actualizar_autenticacion(self, id,estado, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.autenticacion = estado
        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion="Cambio estado de autenticacion en 2 pasos", fecha=fecha,
                     tabla="cb_usuarios_usuario", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()
        return x


class ModuloManager:

    def __init__(self, db):
        self.db = db

    def list_all(self,Usuario):

        if Usuario.username == "admin":
            x = self.db.query(Modulo).filter(Modulo.fkmodulo == None).all()
        else:
            x = Usuario.rol.modulos

        return x



