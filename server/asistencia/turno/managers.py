from ...operaciones.bitacora.managers import *
from server.common.managers import SuperManager
from .models import *
from ..horario.models import *
import requests

import string
import random
import hashlib

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side


class HoraManager(SuperManager):

    def __init__(self, db):
        super().__init__(Hora, db)


class DiaManager(SuperManager):

    def __init__(self, db):
        super().__init__(Dia, db)

    def litar_todo(self):
        return self.db.query(self.entity)

    def tabla_dia(self):
        list = {}
        c = 0
        objeto = self.db.query(self.entity).filter(self.entity.enabled == True).all()

        for x in objeto:
            aux = ''

            for hr in x.hora:
                aux += ' | ' + hr.entrada.strftime("%H:%M") + ' - ' + hr.salida.strftime("%H:%M")

            aux.replace(' | ', '', 1)

            semanaldetalle = self.db.query(Semanaldetalle).filter(Semanaldetalle.fkdia == x.id).first()

            if semanaldetalle:
                eliminar = False
            else:
                eliminar = True



            list[c] = dict(id=x.id,eliminar=eliminar, codigo=x.codigo, nombre=x.nombre, hora=aux)
            c = c + 1

        return list

    def turno_excel(self, ):
        cname = "Turnos.xlsx"

        turnos = self.db.query(self.entity).order_by(self.entity.id.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'turno'

        indice = 1

        ws['A' + str(indice)] = 'ID'
        ws['B' + str(indice)] = 'CODIGO'
        ws['C' + str(indice)] = 'NOMBRE'
        ws['D' + str(indice)] = 'NORMAL'
        ws['E' + str(indice)] = 'ENTRADA'
        ws['F' + str(indice)] = 'SALIDA'

        for i in turnos:
            indice = indice + 1

            ws['A' + str(indice)] = i.id
            ws['B' + str(indice)] = i.codigo
            ws['C' + str(indice)] = i.nombre
            ws['D' + str(indice)] = i.normal

            for h in i.hora:

                ws['E' + str(indice)] = h.entrada
                ws['F' + str(indice)] = h.salida
                indice = indice + 1

        wb.save("server/common/resources/downloads/" + cname)
        return cname

    def importar_excel(self, cname, user, ip):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['ID', 'CODIGO', 'NOMBRE', 'NORMAL', 'ENTRADA', 'SALIDA']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                list = []
                for row in ws.iter_rows(min_row=2):

                    if row[indices['ID']].value is not None and row[indices['ENTRADA']].value is not None:

                        query = self.db.query(self.entity).filter(
                            self.entity.codigo == row[indices['CODIGO']].value).first()

                        if not query:


                            turno = Dia(codigo=row[indices['CODIGO']].value, nombre=row[indices['NOMBRE']].value,
                                        normal=row[indices['NORMAL']].value,
                                        hora=[])

                            list.append(
                                Hora(entrada=row[indices['ENTRADA']].value, salida=row[indices['SALIDA']].value))



                    else:
                        if row[indices['ENTRADA']].value is not None:
                            list.append(
                                Hora(entrada=row[indices['ENTRADA']].value, salida=row[indices['SALIDA']].value))
                        else:
                            turno.hora = list

                            self.db.merge(turno)
                            self.db.flush()
                            list = []

                turno.hora = list

                self.db.merge(turno)
                self.db.flush()
                list = []

                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except Exception as e:
            self.db.rollback()
            if 'UNIQUE constraint' in str(e):
                return {'message': 'duplicado', 'success': False}
            if 'UNIQUE constraint failed' in str(e):
                return {'message': 'codigo duplicado', 'success': False}
            return {'message': str(e), 'success': False}

    def sms(self,telefono,texto):
        destinations = telefono
        message = texto
        senderId = ""
        debug = True

        DiaManager(self.db).altiriaSms(destinations, message, senderId, debug)

        return texto

    def altiriaSms(self,destinations, message, senderId, debug):
        if debug:
            print('Enter altiriaSms: ' + destinations + ', message: ' + message + ', senderId: ' + senderId)

            try:
                # Se crea la lista de parámetros a enviar en la petición POST
                # XX, YY y ZZ se corresponden con los valores de identificación del usuario en el sistema.
                payload = [
                    ('cmd', 'sendsms'),
                    ('domainId', 'demopr'),
                    ('login', 'bvargas@cloudbit.com.bo'),
                    ('passwd', '7fcaxmxb'),
                    # No es posible utilizar el remitente en América pero sí en España y Europa
                    ('senderId', senderId),
                    ('msg', message)
                ]

                # add destinations
                for destination in destinations.split(","):
                    payload.append(('dest', destination))

                # Se fija la codificacion de caracteres de la peticion POST
                contentType = {'Content-Type': 'application/x-www-form-urlencoded;charset=utf-8'}

                # Se fija la URL sobre la que enviar la petición POST
                url = 'http://www.altiria.net/api/http'

                # Se envía la petición y se recupera la respuesta
                r = requests.post(url,
                                  data=payload,
                                  headers=contentType,
                                  # Se fija el tiempo máximo de espera para conectar con el servidor (5 segundos)
                                  # Se fija el tiempo máximo de espera de la respuesta del servidor (60 segundos)
                                  timeout=(5, 60))  # timeout(timeout_connect, timeout_read)

                if debug:
                    if str(r.status_code) != '200':  # Error en la respuesta del servidor
                        print('ERROR GENERAL: ' + str(r.status_code))

                    else:  # Se procesa la respuesta
                        print('Código de estado HTTP: ' + str(r.status_code))

                        if (r.text).find("ERROR errNum:"):
                            print('Error de Altiria: ' + r.text)

                        else:
                            print('Cuerpo de la respuesta: \n' + r.text)

                return r.text

            except requests.ConnectTimeout:
                print("Tiempo de conexión agotado")


            except requests.ReadTimeout:
                print("Tiempo de respuesta agotado")


            except Exception as ex:
                print("Error interno: " + str(ex))


    #print('The function altiriaSms returns: \n' + altiriaSms('346xxxxxxxx,346yyyyyyyy', 'Mesaje de prueba', '', True))


    # No es posible utilizar el remitente en América pero sí en España y Europa
    # Utilizar esta llamada solo si se cuenta con un remitente autorizado por Altiria
    # print 'The function altiriaSms returns: \n'+altiriaSms('346xxxxxxxx,346yyyyyyyy','Mesaje de prueba', 'remitente', True)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()
        i = 0
        for x in objeto.hora:
            objeto.hora[i].entrada = datetime.strptime('01/01/2000 '+objeto.hora[i].entrada, '%d/%m/%Y %H:%M')
            objeto.hora[i].salida = datetime.strptime('01/01/2000 '+objeto.hora[i].salida, '%d/%m/%Y %H:%M')
            i = i + 1
        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registró un turno.", fecha=fecha, tabla="cb_asistencia_dia", identificador=a.id)
        super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        i = 0
        for x in objeto.hora:
            objeto.hora[i].entrada = datetime.strptime('01/01/2000 '+objeto.hora[i].entrada, '%d/%m/%Y %H:%M')
            objeto.hora[i].salida = datetime.strptime('01/01/2000 '+objeto.hora[i].salida, '%d/%m/%Y %H:%M')
            i = i + 1

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modificó un turno.", fecha=fecha,tabla="cb_asistencia_dia", identificador=a.id)
        super().insert(b)
        return a

    def eliminar(self, id, estado, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.enabled = estado
        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion="Eliminó un turno.", fecha=fecha, tabla="cb_asistencia_dia", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()
        return x

