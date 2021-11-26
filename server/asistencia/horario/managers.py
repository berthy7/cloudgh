from ...operaciones.bitacora.managers import *
from server.common.managers import SuperManager
from .models import *

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from ..turno.models import *
from ..asignacion.models import *


class SemanalManager(SuperManager):

    def __init__(self, db):
        super().__init__(Semanal, db)


    def eliminar(self, id, estado, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.enabled = estado
        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion="Eliminó Semanal.", fecha=fecha, tabla = "asistencia_semanal", identificador = id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()
        return x


    def tabla_semanal(self):
        list = {}
        c = 0
        objeto = self.db.query(self.entity).filter(self.entity.enabled == True).all()

        for x in objeto:
            list_libre = {}
            list_libre[0]= dict(hora="Libre")
            lunes = list_libre
            martes = list_libre
            miercoles = list_libre
            jueves = list_libre
            viernes = list_libre
            sabado = list_libre
            domingo = list_libre

            periodo = self.db.query(Periodo).filter(Periodo.fksemanal == x.id).filter(Periodo.enabled == True).first()

            if periodo:
                eliminar = False
            else:
                eliminar = True

            for detalle in x.semanaldetalle:
                list_hora = {}
                ch = 0

                print(x.id)
                for horario in detalle.dia.hora:
                    aux = horario.entrada.strftime("%H:%M") + " - " + horario.salida.strftime("%H:%M")

                    list_hora[ch]= dict(hora=aux)
                    ch = ch + 1

                if detalle.lunes:
                    lunes= list_hora
                if detalle.martes:
                    martes=list_hora
                if detalle.miercoles:
                    miercoles=list_hora
                if detalle.jueves:
                    jueves=list_hora
                if detalle.viernes:
                    viernes=list_hora
                if detalle.sabado:
                    sabado=list_hora
                if detalle.domingo:
                    domingo=list_hora

            list[c] = dict(id=x.id,eliminar=eliminar, nombre=x.nombre,lunes=lunes,martes=martes,miercoles=miercoles,jueves=jueves,viernes=viernes,sabado=sabado,domingo=domingo)
            c = c + 1

        return list

    def listar_todo(self):
        return self.db.query(self.entity)

    def horario_excel(self, ):
        cname = "Horarios.xlsx"

        semanal = self.db.query(self.entity).order_by(self.entity.id.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'a'

        indice = 1

        ws['A' + str(indice)] = 'ID'
        ws['B' + str(indice)] = 'CODIGO'
        ws['C' + str(indice)] = 'NOMBRE'
        ws['D' + str(indice)] = 'TURNO'
        ws['E' + str(indice)] = 'LUNES'
        ws['F' + str(indice)] = 'MARTES'
        ws['G' + str(indice)] = 'MIERCOLES'
        ws['H' + str(indice)] = 'JUEVES'
        ws['I' + str(indice)] = 'VIERNES'
        ws['J' + str(indice)] = 'SABADO'
        ws['k' + str(indice)] = 'DOMINGO'

        for i in semanal:
            indice = indice + 1

            ws['A' + str(indice)] = i.id
            ws['B' + str(indice)] = i.codigo
            ws['C' + str(indice)] = i.nombre

            for h in i.semanaldetalle:

                turno = ""

                if h.fkdia:
                    turno = h.dia.codigo

                ws['D' + str(indice)] = turno
                ws['E' + str(indice)] = h.lunes
                ws['F' + str(indice)] = h.martes
                ws['G' + str(indice)] = h.miercoles
                ws['H' + str(indice)] = h.jueves
                ws['I' + str(indice)] = h.viernes
                ws['J' + str(indice)] = h.sabado
                ws['K' + str(indice)] = h.domingo
                indice = indice + 1

        wb.save("server/common/resources/downloads/" + cname)
        return cname

    def importar_excel(self, cname, user, ip):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['ID', 'CODIGO', 'NOMBRE', 'TURNO', 'LUNES','MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO', 'DOMINGO']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                list = []
                for row in ws.iter_rows(min_row=2):

                    if row[indices['ID']].value is not None and row[indices['TURNO']].value is not None:

                        query = self.db.query(self.entity).filter(
                            self.entity.codigo == row[indices['NOMBRE']].value).first()

                        if not query:


                            horario = Semanal(codigo=row[indices['CODIGO']].value, nombre=row[indices['NOMBRE']].value,
                                            semanaldetalle=[])

                            fkdia = None

                            query_dia = self.db.query(Dia).filter(
                                Dia.codigo == row[indices['TURNO']].value).first()

                            if query_dia:
                                fkdia = query_dia.id



                            list.append(
                                Semanaldetalle(lunes=row[indices['LUNES']].value, martes=row[indices['MARTES']].value,
                                     miercoles=row[indices['MIERCOLES']].value,
                                     jueves=row[indices['JUEVES']].value,
                                     viernes=row[indices['VIERNES']].value,
                                     sabado=row[indices['SABADO']].value,
                                     domingo=row[indices['DOMINGO']].value,
                                     fkdia=fkdia))



                    else:
                        if row[indices['TURNO']].value is not None:
                            fkdia = None

                            query_dia = self.db.query(Dia).filter(
                                Dia.codigo == row[indices['TURNO']].value).first()

                            if query_dia:
                                fkdia = query_dia.id


                            list.append(
                                Semanaldetalle(lunes=row[indices['LUNES']].value, martes=row[indices['MARTES']].value,
                                               miercoles=row[indices['MIERCOLES']].value,
                                               jueves=row[indices['JUEVES']].value,
                                               viernes=row[indices['VIERNES']].value,
                                               sabado=row[indices['SABADO']].value,
                                               domingo=row[indices['DOMINGO']].value,
                                               fkdia=fkdia))
                        else:
                            horario.semanaldetalle = list

                            self.db.merge(horario)
                            self.db.flush()
                            list = []

                horario.semanaldetalle = list

                self.db.merge(horario)
                self.db.flush()
                list = []

                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except Exception as e:
            self.db.rollback()
            if 'UNIQUE constraint' in str(e):
                return {'message': 'duplicado', 'success': False}
            if 'UNIQUE constraint failed' in str(e):
                return {'message': 'codigo duplicado', 'success': False}
            return {'message': str(e), 'success': False}
