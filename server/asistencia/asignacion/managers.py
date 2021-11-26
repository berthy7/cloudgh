from ...operaciones.bitacora.managers import *
from server.common.managers import SuperManager
from .models import *

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from server.asistencia.horario.models import *
from server.personal.persona.models import *

class PeriodoManager(SuperManager):

    def __init__(self, db):
        super().__init__(Periodo, db)

    def list_all(self):
        return dict(objects=self.db.query(self.entity).filter(self.entity.enabled))

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.enabled == True).all()



    def eliminar(self, id, estado, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.enabled = estado
        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion="Eliminó Asignacion.", fecha=fecha, tabla = "asistencia_asignacion", identificador = id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()
        return x

    def periodo_excel(self, ):
        cname = "Asignacion.xlsx"

        semanal = self.db.query(self.entity).order_by(self.entity.id.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'a'

        indice = 1

        ws['A' + str(indice)] = 'ID'
        ws['B' + str(indice)] = 'SEMANAL'
        ws['C' + str(indice)] = 'PERSONA'

        for i in semanal:
            indice = indice + 1

            ws['A' + str(indice)] = i.id
            ws['B' + str(indice)] = i.semanal.nombre

            for a in i.asignacion:

                persona = ""

                if a.fkpersona:
                    persona = a.persona.empleado[0].codigo


                ws['C' + str(indice)] = persona

                indice = indice + 1

        wb.save("server/common/resources/downloads/" + cname)
        return cname

    def importar_excel(self, cname, user, ip):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['ID', 'SEMANAL', 'PERSONA']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                list = []
                for row in ws.iter_rows(min_row=2):

                    if row[indices['ID']].value is not None and row[indices['PERSONA']].value is not None:

                        query_semanal = self.db.query(Semanal).filter(
                            Semanal.nombre == row[indices['SEMANAL']].value).first()

                        periodo = Periodo(fksemanal=query_semanal.id,asignacion=[])

                        fkpersona = None

                        query_persona = self.db.query(Persona).join(Empleado).filter(
                            Empleado.codigo == row[indices['PERSONA']].value).first()

                        if query_persona:
                            fkpersona = query_persona.id



                        list.append(Asignacion(fkpersona=fkpersona))



                    else:
                        if row[indices['PERSONA']].value is not None:
                            fkpersona = None

                            query_persona = self.db.query(Persona).join(Empleado).filter(
                                Empleado.codigo == row[indices['PERSONA']].value).first()

                            if query_persona:
                                fkpersona = query_persona.id

                            list.append(
                                Asignacion(fkpersona=fkpersona))
                        else:
                            periodo.asignacion = list

                            self.db.merge(periodo)
                            self.db.flush()
                            list = []

                periodo.asignacion = list

                self.db.merge(periodo)
                self.db.flush()
                list = []

                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except Exception as e:
            self.db.rollback()
            if 'UNIQUE constraint' in str(e):
                return {'message': 'duplicado', 'success': False}
            if 'UNIQUE constraint failed' in str(e):
                return {'message': 'codigo duplicado', 'success': False}
            return {'message': str(e), 'success': False}



class AsignacionManager(SuperManager):

    def __init__(self, db):
        super().__init__(Asignacion,db)

    def obtener_horarios(self, id):
        objeto = self.db.query(Asignacion).filter(Asignacion.fkpersona == id).first()
        horario_lunes = "------"
        horario_martes = "-----"
        horario_miercoles = "-----"
        horario_jueves = "-----"
        horario_viernes = "-----"
        horario_sabado = "-----"
        horario_domingo = "-----"
        vectotr_horarios = []

        for horarios in objeto.periodo.semanal.semanaldetalle:
            if horarios.lunes:
                horario_lunes = ""
                for hora in horarios.dia.hora:
                    horario_lunes += "<p>" + str(hora.entrada.strftime("%H:%M")) + " - " + str(hora.salida.strftime("%H:%M")) + "</p>"

            if horarios.martes:
                horario_martes = ""
                for hora in horarios.dia.hora:
                    horario_martes += "<p>" + str(hora.entrada.strftime("%H:%M")) + " - " + str(hora.salida.strftime("%H:%M")) + "</p>"

            if horarios.miercoles:
                horario_miercoles = ""
                for hora in horarios.dia.hora:
                    horario_miercoles += "<p>" + str(hora.entrada.strftime("%H:%M")) + " - " + str(hora.salida.strftime("%H:%M")) + "</p>"

            if horarios.jueves:
                horario_jueves = ""
                for hora in horarios.dia.hora:
                    horario_jueves += "<p>" + str(hora.entrada.strftime("%H:%M")) + " - " + str(hora.salida.strftime("%H:%M")) + "</p>"

            if horarios.viernes:
                horario_viernes = ""
                for hora in horarios.dia.hora:
                    horario_viernes += "<p>" + str(hora.entrada.strftime("%H:%M")) + " - " + str(hora.salida.strftime("%H:%M")) + "</p>"

            if horarios.sabado:
                horario_sabado = ""
                for hora in horarios.dia.hora:
                    horario_sabado += "<p>" + str(hora.entrada.strftime("%H:%M")) + " - " + str(hora.salida.strftime("%H:%M")) + "</p>"

            if horarios.domingo:
                horario_domingo = ""
                for hora in horarios.dia.hora:
                    horario_domingo += "<p>" + str(hora.entrada.strftime("%H:%M")) + " - " + str(hora.salida.strftime("%H:%M")) + "</p>"

        vectotr_horarios.append(horario_lunes)
        vectotr_horarios.append(horario_martes)
        vectotr_horarios.append(horario_miercoles)
        vectotr_horarios.append(horario_jueves)
        vectotr_horarios.append(horario_viernes)
        vectotr_horarios.append(horario_sabado)
        vectotr_horarios.append(horario_domingo)

        return vectotr_horarios

    def obtener_x_persona(self, fkpersona):
        objeto = self.db.query(Asignacion).filter(Asignacion.fkpersona == fkpersona).first()


        return objeto
