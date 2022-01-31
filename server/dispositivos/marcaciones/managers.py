from server.common.managers import SuperManager
from ...operaciones.bitacora.managers import BitacoraManager
from ...operaciones.bitacora.models import *
from ...personal.persona.models import *
from ...usuarios.ajustes.models import *
from .models import *
from sqlalchemy.sql import func
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
import csv

import pytz


class MarcacionesManager(SuperManager):

    def __init__(self, db):
        super().__init__(Marcaciones, db)

    def ws_insertRegistros(self,marcaciones):
        for marcacion in marcaciones['marcaciones']:

            marcacion[6] = datetime.strptime(marcacion[6], '%d/%m/%Y %H:%M:%S')

            print("llegaron marcaciones: "+str(marcacion[6]))
            respuesta = self.db.query(self.entity).filter(self.entity.evento == marcacion[4]).filter(self.entity.time == marcacion[6]).filter(self.entity.tarjeta == marcacion[0]).filter(self.entity.fkdispositivo == marcaciones['iddispositivo']).first()

            if not respuesta:
                print("registro marcacion")

                object = Marcaciones(tarjeta=marcacion[0],codigo=marcacion[1],verificado=marcacion[2],puerta=marcacion[3],evento=marcacion[4],estado=marcacion[5],time=marcacion[6],fkdispositivo=marcaciones['iddispositivo'])

                self.db.add(object)

        self.db.commit()
        self.db.close()

    # def ws_insertRegistros_biometricos(self, marcaciones):
    #     for marcacion in marcaciones['marcaciones']:
    #
    #         marcacion[1] = datetime.strptime(marcacion[1], '%Y-%m-%d %H:%M:%S')
    #
    #         # print("llegaron marcaciones: " + str(marcacion[1]))
    #
    #         respuesta = self.db.query(self.entity).filter(self.entity.time == marcacion[1]).filter(self.entity.codigo == marcacion[0]).filter(
    #             self.entity.fkdispositivo == marcaciones['iddispositivo']).first()
    #
    #         if not respuesta:
    #             print("registro marcacion")
    #
    #             object = Marcaciones(codigo=marcacion[0], time=marcacion[1],
    #                                  fkdispositivo=marcaciones['iddispositivo'])
    #             AsistenciaManager(self.db).insertar_marcaciones(marcacion[1], marcacion[0])
    #
    #             self.db.add(object)
    #
    #     self.db.commit()
    #     self.db.close()

    def get_all(self):
        return self.db.query(self.entity).filter(self.entity.enabled == True).order_by(self.entity.nombre.asc()).all()

    def list_all(self):
        return dict(
            objects=self.db.query(self.entity).filter(self.entity.enabled == True))

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.enabled == True)

    def listar_por_dia(self):
        list = {}
        c = 0

        fecha_zona = datetime.now(pytz.timezone('America/La_Paz'))
        fecha = fecha_zona
        fechahoy = str(fecha.day) + "/" + str(fecha.month) + "/" + str(fecha.year)
        fechahoy = datetime.strptime(fechahoy, '%d/%m/%Y')

        ajuste = self.db.query(Ajustes).filter(Ajustes.enabled == True).first()

        if ajuste.oracle:
            # version oracle
            objeto = self.db.query(self.entity).filter(func.to_date(self.entity.time).between(fechahoy, fechahoy)).all()
        else:
            # version postgres
            objeto = self.db.query(self.entity).filter(func.date(self.entity.time).between(fechahoy, fechahoy)).all()

        for x in objeto:
            try:
                empleado = self.db.query(Empleado).join(Persona).filter(Empleado.codigo == x.codigo).one()
                nombrepersona = empleado.persona.fullname
            except Exception as ex:
                nombrepersona= "----"

            list[c] = dict(id=x.id,codigo=x.codigo, nombre=nombrepersona, fecha=x.time.strftime("%d/%m/%Y"), hora=x.time.strftime("%H:%M:%S"), dispositivo=x.dispositivo.descripcion)
            c = c + 1

        return list

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip_maquina, accion="Registro Empresa.",
                     fecha=fecha, tabla="rrhh_empresa", identificador=a.id)
        super().insert(b)
        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Empresa.",
                     fecha=fecha, tabla="rrhh_empresa", identificador=a.id)
        super().insert(b)
        return a

    def delete(self, id, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.enabled = False
        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion="Eliminó Empresa.", fecha=fecha,
                     tabla="rrhh_empresa", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()

    def filtrar(self, fechainicio, fechafin):
        list = {}
        c = 0

        ajuste = self.db.query(Ajustes).filter(Ajustes.enabled == True).first()

        if ajuste.oracle:
            # version oracle
            objeto = self.db.query(self.entity).filter(func.to_date(self.entity.time).between(fechainicio, fechafin)).order_by(self.entity.id.desc()).all()
        else:
            # version postgres
            objeto = self.db.query(self.entity).filter(func.date(self.entity.time).between(fechainicio, fechafin)).order_by(self.entity.id.desc()).all()

        for x in objeto:
            try:
                empleado = self.db.query(Empleado).join(Persona).filter(Empleado.codigo == x.codigo).one()
                nombrepersona = empleado.persona.fullname
            except Exception as ex:
                nombrepersona= "----"

            if x.fkdispositivo:
                dispositivo = x.dispositivo.descripcion
            else:
                dispositivo = "----"


            list[c] = dict(id=x.id,codigo=x.codigo, nombre=nombrepersona, fecha=x.time.strftime("%d/%m/%Y"), hora=x.time.strftime("%H:%M:%S"), dispositivo=dispositivo)
            c = c + 1

        return list

    def obtener_marcaciones(self,codigo, fechainicio, fechafin):

        ajuste = self.db.query(Ajustes).filter(Ajustes.enabled == True).first()

        if ajuste.oracle:
            # version oracle
            objeto = self.db.query(self.entity).filter(self.entity.codigo == codigo).filter(func.to_date(self.entity.time).between(fechainicio, fechafin)).order_by(self.entity.id.asc()).all()

        else:
            # version postgres
            objeto = self.db.query(self.entity).filter(self.entity.codigo == codigo).filter(
                func.date(self.entity.time).between(fechainicio, fechafin)).order_by(self.entity.id.asc()).all()

        return objeto

    def importar_excel(self, cname):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['ID', 'Fecha']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                cont = 1
                for row in ws.iter_rows(min_row=2):

                    date_time = datetime.strptime(str(row[indices['Fecha']].value), '%d/%m/%Y %H:%M')

                    marc = Marcaciones(codigo=int(row[indices['ID']].value), time=date_time)

                    self.db.add(marc)
                    # self.db.flush()
                    cont = cont + 1

                print("inicio commit")
                self.db.commit()
                print("fin commit")
                return {'message': 'Marcaciones Importadas Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except Exception as e:
            print(e)
            return {'message': 'Error', 'success': False}

    def importar_txt(self, cname):
        """
        Parameters
        ----------
        cname:string

        Returns
        -------
        Devuelve la confirmacion de importacion.
        """
        with open("server/common/resources/uploads/" + cname, 'rt') as f:

            reader = csv.reader(f, delimiter=',')
            cont = 1
            try:
                for line in reader:
                    legajo_str = line[0]
                    if cont == 1:
                        legajo = legajo_str[3:len(legajo_str)]
                    else:
                        legajo = legajo_str

                    time_srt = line[1].split(' ')
                    fecha = time_srt[0]
                    hora_str = time_srt[3]
                    hora = hora_str[0:8]

                    id = line[2]

                    print(legajo+" "+fecha +" "+ hora+" "+id)

                    date_time = datetime.strptime(str(fecha) + " " + str(hora), '%Y-%m-%d %H:%M:%S')

                    marc = Marcaciones(codigo=int(legajo),time=date_time)

                    self.db.add(marc)
                    # self.db.flush()
                    cont = cont + 1

                    # fecha_ingreso = fecha_final = None
                    # fecha_ingreso = datetime.strptime(line[4], '%d/%m/%Y')
                    # if line[5] is not '':
                    #     fecha_final = datetime.strptime(line[5], '%d/%m/%Y')
                    # empl = Empleado(codigo=line[3])
                    #
                    # empl.retiro.append(ret)
                    # persona = Persona(
                    #     nombres=nombres,
                    #     apellido_paterno=words[0],
                    #     apellido_materno=words[1].replace(",", ""),
                    #     sexo=line[1].title(),
                    #     dni=line[2],
                    # )
                    # persona.empleado.append(empl)
                    # self.db.add(persona)
                    # self.db.commit()
                try:
                    self.db.commit()
                except Exception as e:
                    print(e)

                return {'message': 'Importado Todos Correctamente.', 'success': True}
            except Exception as e:
                print(e)
                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}

    def importar_txt_monterrey(self, cname):
        """
        Parameters
        ----------
        cname:string

        Returns
        -------
        Devuelve la confirmacion de importacion.
        """
        with open("server/common/resources/uploads/" + cname, 'rt') as f:

            reader = csv.reader(f, delimiter=',')
            cont = 1
            try:
                for line in reader:
                    id = line[0]
                    codigo = line[1]

                    time_srt = line[2].split(' ')
                    fecha = time_srt[0]
                    hora = time_srt[1]
                    # hora = hora_str[0:8]

                    sucursal = line[3]

                    print(codigo+" "+fecha +" "+ hora+" "+id)

                    date_time = datetime.strptime(str(fecha) + " " + str(hora), '%d/%m/%Y %H:%M:%S')

                    marc = Marcaciones(codigo=int(codigo),time=date_time,sucursal=sucursal)

                    self.db.add(marc)
                    # self.db.flush()
                    cont = cont + 1


                try:
                    self.db.commit()
                except Exception as e:
                    print(e)

                return {'message': 'Importado Todos Correctamente.', 'success': True}
            except Exception as e:
                print(e)
                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}