from .models import *
from ...configuraciones.cargo.managers import *
from ...configuraciones.ciudad.managers import *
from ...configuraciones.empresa.models import *
from ..persona.models import *

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font


class OrganigramaManager(SuperManager):
    def __init__(self, db):
        super().__init__(Organigrama, db)

    def list_all(self):
        return dict(objects=self.db.query(Organigrama).filter(Organigrama.enabled is True))

    def organigrama_excel(self, ):
        cname = "Organigrama.xlsx"

        organigrama = self.db.query(self.entity).order_by(self.entity.id.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'organigrama'

        indice = 1

        ws['A' + str(indice)] = 'ID'
        ws['B' + str(indice)] = 'CARGO'
        ws['C' + str(indice)] = 'CODIGO'
        ws['D' + str(indice)] = 'SUPERIOR'
        ws['E' + str(indice)] = 'POSICION'
        ws['F' + str(indice)] = 'GERENCIA'
        ws['G' + str(indice)] = 'JEFATURA'
        ws['H' + str(indice)] = 'ESTADO'

        for i in organigrama:

            if i.fkcargo:
                indice = indice + 1
                cargo = self.db.query(Cargo).filter(Cargo.id == i.fkcargo).first()

                if cargo:
                    NombreCargo = cargo.nombre
                else:
                    NombreCargo = ""

                persona = self.db.query(Persona).filter(Persona.id == i.fkpersona).first()

                if persona:
                    CodigoPersona = persona.empleado[0].codigo
                else:
                    CodigoPersona = ""


                if persona:
                    CodigoPersona = persona.empleado[0].codigo
                else:
                    CodigoPersona = ""


                ws['A' + str(indice)] = i.id
                ws['B' + str(indice)] = NombreCargo
                ws['C' + str(indice)] = CodigoPersona
                ws['D' + str(indice)] = i.fkpadre
                ws['E' + str(indice)] = i.siguiente
                ws['F' + str(indice)] = i.gerencia
                ws['G' + str(indice)] = i.jefatura
                ws['H' + str(indice)] = i.enabled

        wb.save("server/common/resources/downloads/" + cname)
        return cname

    def importar_excel(self, cname, user, ip):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['ID', 'CARGO', 'CODIGO', 'SUPERIOR', 'POSICION', 'GERENCIA', 'JEFATURA', 'ESTADO']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                list = []

                padre = self.db.query(Organigrama).order_by(self.entity.id.asc()).first()

                sw = 0

                for row in ws.iter_rows(min_row=2):
                    sw = sw+1

                    if row[indices['ID']].value is not None and row[indices['CARGO']].value is not None  :


                        query_cargo = self.db.query(Cargo).filter(Cargo.nombre == row[indices['CARGO']].value).first()

                        if query_cargo:

                            query_persona = self.db.query(Persona).join(Empleado).filter(
                                Empleado.codigo == row[indices['CODIGO']].value).first()

                            if query_persona:

                                    if sw == 1:

                                        fkpadre = padre.id
                                    else:
                                        fkpadre = row[indices['SUPERIOR']].value


                                    org = Organigrama(
                                                    id=row[indices['ID']].value,
                                                    fkcargo=query_cargo.id,
                                                    fkpersona=query_persona.id,
                                                    fkpadre=fkpadre,
                                                    siguiente=row[indices['POSICION']].value,
                                                    gerencia=row[indices['GERENCIA']].value,
                                                    jefatura=row[indices['JEFATURA']].value,
                                                    enabled=row[indices['ESTADO']].value)

                                    self.db.merge(org)
                                    self.db.flush()


                    else:
                        padre = padre.id

                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except Exception as e:
            self.db.rollback()
            if 'UNIQUE constraint' in str(e):
                return {'message': 'duplicado', 'success': False}
            if 'UNIQUE constraint failed' in str(e):
                return {'message': 'codigo duplicado', 'success': False}
            return {'message': str(e), 'success': False}

    def list_son(self):
        listaID=self.list_all()
        listaHijos=list()
        listaFK=self.list_all()
        fks=list()
        for lid in listaID:
            fks.append(lid.fkpadre)
        for lfk in listaFK:
            if lfk.id not in fks:
                listaHijos.append(lfk)
        return listaHijos

    def get_brother(self, fkpadre, id):
        return dict(contador_bro=self.db.query(self.entity).filter(self.entity.enabled == True).\
            filter(self.entity.fkpadre == fkpadre).count(), contador_son=self.db.query(self.entity).filter(self.entity.enabled == True).\
            filter(self.entity.fkpadre == id).count())

    def get_all(self):
        query = self.db.query(self.entity).filter(self.entity.enabled == True).filter(self.entity.fkpadre == None).first()
        s = self.order_tree(query)
        return s

    def order_tree(self, tree):
        tree2 = tree
        return self.order_tree_pri(tree, dict(), tree2)

    def order_tree_pri(self, tree, item, item_tree):
        if tree:
            # rs = tree.pop()
            if tree.fkcargo:
                cargo = tree.cargo.nombre
            else:
                empresa = self.db.query(Empresa).filter(Empresa.enabled == True).first()
                if empresa:
                    cargo = empresa.nombre
                else:
                    cargo = "Por Definir"

            if tree.fkpersona:
                persona = tree.persona.fullname
            else:
                persona = ""

            item = dict(id=tree.id, name=cargo, title=persona, fk_parent='', children=[])
            for h in tree.hijos:

                h_copia = h
                item_hijos = self.order_tree_pri(h, item['children'],  h_copia)
                item['children'].append(item_hijos)

            return item
        else:
            return item

    def checked_object_position(self,pos,fkpadre):
        return self.db.query(self.entity).filter(self.entity.fkpadre==fkpadre).filter(self.entity.siguiente==pos).first()

    def update(self, object):
        pos = self.checked_object_position(object.siguiente,object.fkpadre)
        obj = self.obtain(object.id)
        if pos is not None:
            pos.siguiente = obj.siguiente
            super().update(pos)
        return super().update(object)


    def obtener_superior(self,fkpersona):

        # Obtener Superior del Organigrama
        x = self.db.query(self.entity).filter(self.entity.fkpersona==fkpersona).first()
        
        #Si no hay Superior de 'x', devolver None
        if(x is None):
            return None
        # Obtener Padre
        padre = self.db.query(self.entity).filter(self.entity.id == x.fkpadre).first()



        return padre.fkpersona

    def obtener_autorizacion_aprobacion(self, fkpersona):

        # Obtener Superior del Organigrama
        superiores = dict(fkautorizacion=None, fkaprobacion=None, estadoautorizacion="Pendiente",
                          estadoaprobacion="Pendiente")


        Idpersona = fkpersona

        while True:
            organi = self.db.query(self.entity).filter(self.entity.fkpersona == Idpersona).first()

            # Obtener fkautorizacion


            if organi.jefatura:
                superiores['fkautorizacion'] = organi.fkpersona

                padre = self.db.query(self.entity).filter(self.entity.id == organi.fkpadre).first()
                Idpersona = padre.fkpersona


            elif organi.gerencia:
                superiores['fkaprobacion'] = organi.fkpersona

                break
            else :
                padre = self.db.query(self.entity).filter(self.entity.id == organi.fkpadre).first()
                Idpersona = padre.fkpersona

        return superiores


    def obtener_autorizacion_aprobacion2(self, fkpersona):
        persona = self.db.query(Persona).filter(Persona.id == fkpersona).first()

        # Obtener Superior del Organigrama
        x = self.db.query(self.entity).filter(self.entity.fkpersona == fkpersona).first()

        # Si no hay Superior de 'x', devolver None
        if (x is None):
            superiores = dict(fkautorizacion=None, fkaprobacion=None,estadoautorizacion="Pendiente",estadoaprobacion="Pendiente")

            return superiores

        if persona.empleado[0].autorizacion and persona.empleado[0].aprobacion:

            # Obtener fkautorizacion
            autorizacion = self.db.query(self.entity).filter(self.entity.id == x.fkpadre).first()

            if autorizacion:
                # Obtener fkaprobacion
                aprobacion = self.db.query(self.entity).filter(self.entity.id == autorizacion.fkpadre).first()
            else:
                aprobacion = None

            superiores = dict(fkautorizacion=autorizacion.fkpersona,fkaprobacion=aprobacion.fkpersona,estadoautorizacion="Pendiente",estadoaprobacion="Pendiente")


        elif persona.empleado[0].autorizacion:

            # Obtener fkautorizacion
            autorizacion = self.db.query(self.entity).filter(self.entity.id == x.fkpadre).first()


            superiores = dict(fkautorizacion=autorizacion.fkpersona,fkaprobacion=None,estadoautorizacion="Pendiente",estadoaprobacion="No aplica")

        elif persona.empleado[0].aprobacion:

            # Obtener fkautorizacion
            autorizacion = self.db.query(self.entity).filter(self.entity.id == x.fkpadre).first()

            superiores = dict(fkautorizacion=None, fkaprobacion=autorizacion.fkpersona,estadoautorizacion="No aplica",estadoaprobacion="Pendiente")
        else:
            superiores = dict(fkautorizacion=None, fkaprobacion=None,estadoautorizacion="Pendiente",estadoaprobacion="Pendiente")






        return superiores

    # def get_all(self):
    #     return self.db.query(self.entity).filter(self.entity.enabled == True).all()
    #
    # def list_all(self):
    #     return dict(objects=self.db.query(self.entity).filter(self.entity.enabled == True))
    #
    # def listar_todo(self):
    #     return self.db.query(self.entity).filter(self.entity.enabled == True)
    #
    # def insert(self, objeto):
    #     fecha = BitacoraManager(self.db).fecha_actual()
    #
    #     a = super().insert(objeto)
    #     b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Organigrama.", fecha=fecha,tabla="rrhh_organigrama", identificador=a.id)
    #     super().insert(b)
    #
    #     return a
    #
    # def update(self, objeto):
    #     fecha = BitacoraManager(self.db).fecha_actual()
    #
    #     a = super().update(objeto)
    #     b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Organigrama.", fecha=fecha,tabla="rrhh_organigrama", identificador=a.id)
    #     super().insert(b)
    #     return a
    #
    # def delete(self, id, user, ip):
    #     x = self.db.query(self.entity).filter(self.entity.id == id).one()
    #     x.enabled = False
    #     fecha = BitacoraManager(self.db).fecha_actual()
    #     b = Bitacora(fkusuario=user, ip=ip, accion="Eliminó Organigrama.", fecha=fecha,tabla="rrhh_organigrama", identificador=id)
    #     super().insert(b)
    #     a = self.db.merge(x)
    #     self.db.commit()
    #     return a

