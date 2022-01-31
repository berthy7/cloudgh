from .models import *
from ...configuraciones.empresa.managers import *
from ...configuraciones.gerencia.managers import *
from ...configuraciones.cargo.managers import *
from ...configuraciones.centro_costo.managers import *
from ...configuraciones.pais.managers import *
from ...configuraciones.departamento.managers import *
from ...configuraciones.ciudad.managers import *
from ...configuraciones.sucursal.managers import *

from ..organigrama.models import *

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font


class PersonaManager(SuperManager):
    def __init__(self, db):
        super().__init__(Persona, db)

    def obtener_x_id(self,idpersona):
        return self.db.query(self.entity).filter(self.entity.id == idpersona).first()

    def listar_todos(self):
        return self.db.query(self.entity).all()


    def get_all(self):
        return self.db.query(self.entity).filter(self.entity.enabled == True).all()

    def get_all_per(self):
        return self.db.query(self.entity).order_by(self.entity.enabled.desc()).all()

    def get_all_by_lastname(self):
        return self.db.query(self.entity).filter(self.entity.enabled == True).order_by(self.entity.apellidopaterno.asc()).all()

    def list_all(self):
        return dict(objects=self.db.query(self.entity).join(Empleado).filter(self.entity.enabled == True))

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.enabled == True).order_by(self.entity.apellidopaterno.asc()).all()

    def obtener_x_gerencia(self,idsucursal,idgerencia):
        objeto= self.db.query(Persona).join(Empleado).filter(Persona.enabled == True).filter(Empleado.fksucursal== idsucursal).filter(Empleado.fkgerencia== idgerencia).order_by(Persona.fullname.asc()).all()

        return objeto

    def obtener_x_cargo(self,idcargo):
        objeto= self.db.query(Persona).join(Empleado).filter(Persona.enabled == True).filter(Empleado.fkcargo== idcargo).order_by(Persona.fullname.asc()).all()

        return objeto

    def obtener_correo(self,idpersona):
        objeto= self.db.query(Persona).join(Empleado).filter(Persona.enabled == True).filter(Persona.id== idpersona).first()

        return objeto.empleado[0].email

    def obtener_telefono(self,idpersona):
        objeto= self.db.query(Persona).join(Empleado).filter(Persona.enabled == True).filter(Persona.id== idpersona).first()

        return objeto.telefono

    def obtener_correo_superior(self,idpersona):
        objeto = False
        organi_perso = self.db.query(Organigrama).filter(Organigrama.enabled == True).filter(Organigrama.fkpersona == idpersona).first()

        if organi_perso:
            organi_super = self.db.query(Organigrama).filter(Organigrama.enabled == True).filter(
                Organigrama.id == organi_perso.fkpadre).first()
            if organi_super:

                obj= self.db.query(Persona).join(Empleado).filter(Persona.enabled == True).filter(Persona.id== organi_super.fkpersona).first()
                objeto = obj.empleado[0].email


        return objeto

    def obtener_contrato(self, id):
        list = {}
        c = 0

        for table in self.db.query(Contrato).filter(Contrato.enabled == True).filter(Contrato.fkpersona == id).order_by(Contrato.nroContrato.asc()):
            if table:
                if table.fechaIngreso:
                    fini = table.fechaIngreso.strftime("%d/%m/%Y")
                else:
                    fini = None
                if table.fechaForzado:
                    frtf = table.fechaForzado.strftime("%d/%m/%Y")
                else:
                    frtf = None
                if table.fechaFin:
                    ffin = table.fechaFin.strftime("%d/%m/%Y")
                else:
                    ffin = None

                list[c] = dict(id=table.id, nroContrato=table.nroContrato, persona=table.persona.fullname, tipo=table.tipo,
                                fechaIngreso=fini, fechaForzado=frtf, fechaFin=ffin)
                c = c + 1

        return list

    def obtener_all_contratos(self, id):
        list = {}
        c = 0

        for table in self.db.query(Contrato).filter(Contrato.fkpersona == id).order_by(Contrato.nroContrato.asc()):
            if table:
                if table.fechaIngreso:
                    fini = table.fechaIngreso.strftime("%d/%m/%Y")
                else:
                    fini = None
                if table.fechaForzado:
                    frtf = table.fechaForzado.strftime("%d/%m/%Y")
                else:
                    frtf = None
                if table.fechaFin:
                    ffin = table.fechaFin.strftime("%d/%m/%Y")
                else:
                    ffin = None

                list[c] = dict(id=table.id, nroContrato=table.nroContrato, persona=table.persona.fullname, tipo=table.tipo,
                                fechaIngreso=fini, fechaForzado=frtf, fechaFin=ffin)
                c = c + 1

        return list

    def insert(self, objeto):
        objeto.fechanacimiento = datetime.strptime(objeto.fechanacimiento, '%d/%m/%Y')
        for contra in objeto.contrato:

            if contra.sueldo == "":
                contra.sueldo = None


            if contra.fechaFin:
                contra.fechaFin = datetime.strptime(contra.fechaFin, '%d/%m/%Y')
            else:
                contra.fechaFin = None

            if contra.fechaIngreso:
                contra.fechaIngreso = datetime.strptime(contra.fechaIngreso, '%d/%m/%Y')
            else:
                contra.fechaIngreso = None

            if contra.fechaForzado:
                contra.fechaForzado = datetime.strptime(contra.fechaForzado, '%d/%m/%Y')
            else:
                contra.fechaForzado = None

        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Persona.", fecha=fecha,tabla="rrhh_persona", identificador=a.id)
        super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()
        objeto.fechanacimiento = datetime.strptime(objeto.fechanacimiento, '%d/%m/%Y')
        for contra in objeto.contrato:

            if contra.sueldo == "":
                contra.sueldo = None


            if contra.fechaFin:
                contra.fechaFin = datetime.strptime(contra.fechaFin, '%d/%m/%Y')
            else:
                contra.fechaFin = None

            if contra.fechaIngreso:
                contra.fechaIngreso = datetime.strptime(contra.fechaIngreso, '%d/%m/%Y')
            else:
                contra.fechaIngreso = None

            if contra.fechaForzado:
                contra.fechaForzado = datetime.strptime(contra.fechaForzado, '%d/%m/%Y')
            else:
                contra.fechaForzado = None

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Persona.", fecha=fecha,tabla="rrhh_persona", identificador=a.id)
        super().insert(b)
        return a

    def retiro(self, diccionary):
        x = self.db.query(self.entity).filter(self.entity.id == diccionary['idNombre']).one()
        x.enabled = False
        c = self.db.query(Contrato).filter(Contrato.enabled == True).filter(Contrato.fkpersona == diccionary['idNombre']).first()
        if c:
            c.enabled = False
            c.fechaForzado = datetime.strptime(diccionary['fechaForzado'], '%d/%m/%Y')
            c.descripcion = diccionary['descripcion']

        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=diccionary['user'], ip=diccionary['ip'], accion="Eliminó Persona.", fecha=fecha,tabla="rrhh_persona", identificador=diccionary['idNombre'])
        super().insert(b)
        a = self.db.merge(x)
        self.db.commit()
        return a

    def importar_excel(self, cname):
        try:
            cont = 0
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['PAIS', 'DEPARTAMENTO', 'CIUDAD', 'EMPRESA', 'SUCURSAL', 'SECCION', 'CENTRO DE COSTO', 'CODIGO', 'NOMBRE',
                        'APELLIDO_P', 'APELLIDO_M', 'CI', 'SEXO', 'FECHA_NACIMIENTO', 'CARGO',
                        'CONTRATO_PLAZO', 'FECHA_INGRESO_NOM', 'FECHA_VENCIMIENTO', 'CORREO']
            indices = {cell[0].value: n-1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if cell[0].value in colnames}
            if len(indices) == len(colnames):
                for row in ws.iter_rows(min_row=2):
                    cont = cont +1
                    if row[indices['SECCION']].value is not None and row[indices['CODIGO']].value is not None and \
                                    row[indices['EMPRESA']].value is not None and \
                                    row[indices['NOMBRE']].value is not None and \
                                    row[indices['APELLIDO_P']].value is not None and \
                                    row[indices['CENTRO DE COSTO']].value is not None and \
                                    row[indices['CI']].value is not None and \
                                    row[indices['SEXO']].value is not None and \
                                    row[indices['FECHA_NACIMIENTO']].value is not None and \
                                    row[indices['CARGO']].value is not None and \
                                    row[indices['CONTRATO_PLAZO']].value is not None and \
                                    row[indices['FECHA_INGRESO_NOM']].value is not None and \
                                    row[indices['PAIS']].value is not None and \
                                    row[indices['DEPARTAMENTO']].value is not None and \
                                    row[indices['CIUDAD']].value is not None and \
                                    row[indices['SUCURSAL']].value is not None:
                        query = self.db.query(Empleado).filter(Empleado.codigo == row[indices['CODIGO']].value).filter(Persona.ci == str(row[indices['CI']].value)).all()
                        if not query:

                            fechaNacimiento = datetime.strptime(row[indices['FECHA_NACIMIENTO']].value, '%d/%m/%Y')
                            fechaIngreso = datetime.strptime(row[indices['FECHA_INGRESO_NOM']].value, '%d/%m/%Y')

                            if row[indices['FECHA_VENCIMIENTO']].value == "":
                                fechaFin = None
                            else:
                                fechaFin = datetime.strptime(row[indices['FECHA_VENCIMIENTO']].value, '%d/%m/%Y')

                            empl = Empleado(codigo=row[indices['CODIGO']].value,email=row[indices['CORREO']].value)
                            persona = Persona(nombres=row[indices['NOMBRE']].value,
                                              apellidopaterno=row[indices['APELLIDO_P']].value,
                                              apellidomaterno=row[indices['APELLIDO_M']].value,
                                              sexo=row[indices['SEXO']].value.title(),
                                              ci=row[indices['CI']].value,
                                              fechanacimiento=fechaNacimiento,
                                              )
                            contra = Contrato(
                                nroContrato=row[indices['CODIGO']].value,
                                fechaIngreso=fechaIngreso,
                                fechaFin=fechaFin,
                                tipo=row[indices['CONTRATO_PLAZO']].value)

                            if EmpresaManager(self.db).obtener_x_nombre(row[indices['EMPRESA']].value.title()) is None:
                                empresa = Empresa(nombre=row[indices['EMPRESA']].value.title())
                            else:
                                empresa = EmpresaManager(self.db).obtener_x_nombre(row[indices['EMPRESA']].value.title())

                            if GerenciaManager(self.db).obtener_x_nombre(row[indices['SECCION']].value.title(),
                                                                     row[indices['EMPRESA']].value.title()) is None:
                                gerencia = Gerencia(nombre=row[indices['SECCION']].value.title())
                                gerencia.empresa = empresa
                                empl.gerencia = gerencia
                            else:
                                empl.gerencia = GerenciaManager(self.db).obtener_x_nombre(
                                    row[indices['SECCION']].value.title(), row[indices['EMPRESA']].value.title())

                            if CargoManager(self.db).obtener_x_nombre(row[indices['CARGO']].value.title()) is None:
                                cargo = Cargo(nombre=row[indices['CARGO']].value.title())
                                empl.cargo = cargo
                            else:
                                empl.cargo = CargoManager(self.db).obtener_x_nombre(row[indices['CARGO']].value.title())

                            if Centro_costoManager(self.db).obtener_x_nombre(
                                    row[indices['CENTRO DE COSTO']].value.title()) is None:
                                centro_costo = Centro_costo(nombre=row[indices['CENTRO DE COSTO']].value.title())
                                empl.centro = centro_costo
                            else:
                                empl.centro = Centro_costoManager(self.db).obtener_x_nombre(
                                    row[indices['CENTRO DE COSTO']].value.title())

                            if PaisManager(self.db).obtener_x_nombre(row[indices['PAIS']].value.title()) is None:
                                pais = Pais(nombre=row[indices['PAIS']].value.title())
                                empl.pais = pais
                            else:
                                empl.pais = PaisManager(self.db).obtener_x_nombre(row[indices['PAIS']].value.title())

                            if DepartamentoManager(self.db).obtener_x_nombre(
                                    row[indices['DEPARTAMENTO']].value.title()) is None:
                                departamentos = Departamento(nombre=row[indices['DEPARTAMENTO']].value.title())
                                empl.pais.departamentos.append(departamentos)
                                empl.departamentos = departamentos

                            else:
                                empl.departamentos = DepartamentoManager(self.db).obtener_x_nombre(
                                    row[indices['DEPARTAMENTO']].value.title())

                            if CiudadManager(self.db).obtener_x_nombre(row[indices['CIUDAD']].value.title()) is None:
                                ciudad = Ciudad(nombre=row[indices['CIUDAD']].value.title())
                                empl.departamentos.ciudades.append(ciudad)
                                empl.ciudad = ciudad
                            else:
                                empl.ciudad = CiudadManager(self.db).obtener_x_nombre(row[indices['CIUDAD']].value.title())

                            if SucursalManager(self.db).obtener_x_nombre(row[indices['SUCURSAL']].value.title()) is None:
                                sucursal = Sucursal(nombre=row[indices['SUCURSAL']].value.title())
                                empl.ciudad.sucursales.append(sucursal)
                                empl.sucursal = sucursal
                                empl.sucursal.empresa = empresa
                            else:
                                empl.sucursal = SucursalManager(self.db).obtener_x_nombre(
                                    row[indices['SUCURSAL']].value.title())
                            persona.contrato.append(contra)
                            persona.empleado.append(empl)
                            self.db.add(persona)
                            self.db.flush()
                    else:
                        self.db.rollback()
                        return {'message': 'Hay Columnas vacias', 'success': False}

                self.db.commit()

                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except Exception as e:
            self.db.rollback()

            if 'UNIQUE constraint failed: rrhh_persona.dni' in str(e):
                return {'message': 'CI duplicado', 'success': False}
            if 'UNIQUE constraint failed: rrhh_empleado.codigo' in str(e):
                return {'message': 'codigo de empleado duplicado', 'success': False}
            return {'message': str(e), 'success': False}

    def get_employees_tree2(self):
        query = self.db.query(Empleado.id,
                              Persona.fullname,
                              Persona.sexo,
                              Empresa.id,
                              Empresa.nombre,
                              Sucursal.id,
                              Sucursal.nombre,
                              Gerencia.id,
                              Gerencia.nombre).\
            join(Gerencia, Persona, Sucursal, Empresa).\
            filter(Empleado.enabled==True).all()

        admin = dict()

        for e_id, e_nombre, e_sexo, em_id, em_nombre, s_id, s_nombre, ge_id, ge_nombre in query:
            em = (em_id, em_nombre)
            s = (s_id, s_nombre)
            m = (ge_id, ge_nombre)
            aux = str(e_id)
            html_e = '<li class="dd-item" data-id="' + aux + aux + '"><div class="dd-handle"><input id="' + aux + aux + '" data-id="' + aux + '" data-sex="' + e_sexo + '"type="checkbox" class="module chk-col-deep-purple employee"><label for="' + aux + aux + '">' + e_nombre + '</label></div></li>'
            if em in admin:
                if s[0] is None:
                    admin[em] += html_e
                else:
                    if s in admin[em]:
                        if m[0] is None:
                            admin[em][s] += html_e
                        else:
                            if m in admin[em][s]:
                                admin[em][s][m] += html_e

                            else:
                                admin[em][s][m] = html_e

                    else:
                        if m[0] is None:
                            admin[em][s] += html_e
                        else:
                            admin[em][s] = dict()
                            admin[em][s][m] = html_e

            else:
                if s[0] is None:
                    admin[em] = html_e
                else:
                    if m[0] is None:
                        admin[em] = dict()
                        admin[em][s] = html_e
                    else:
                        admin[em] = dict()
                        admin[em][s] = dict()
                        admin[em][s][m] = html_e

        return admin

    # def get_employees_tree(self):
    #     query = EmpresaManager(self.db).get_all()
    #     admin = dict()
    #
    #     for empresa in query:
    #         em = (empresa.id, empresa.nombre)
    #         admin[em] = dict()
    #         cont_suc = 1
    #         for sucursales in empresa.sucursales:
    #
    #             su = (cont_suc, sucursales.nombre)
    #             cont_suc = cont_suc + 1
    #             admin[em][su] = dict()
    #
    #             list_gerencias = SucursalManager(self.db).obtener_gerencias(empresa.id, sucursales.id)
    #             cont_ger = 1
    #             for gerencias in list_gerencias:
    #                 ger = (cont_ger, list_gerencias[gerencias]['gerencia'])
    #                 cont_ger = cont_ger + 1
    #                 admin[em][su][ger] = dict()
    #
    #                 list_personal = PersonaManager(self.db).obtener_x_gerencia(sucursales.id,list_gerencias[gerencias]['idgerencia'])
    #                 html_e = ""
    #                 for personal in list_personal:
    #                     html = '<li class="dd-item" data-id="' + str(personal.id) + str(
    #                         personal.id) + '"><div class="dd-handle"><input id="' + str(personal.id) + str(
    #                         personal.id) + '" data-id="' + str(personal.id) + '" data-sex="' + str(
    #                         personal.sexo) + '"type="checkbox" class="module chk-col-deep-purple employee"><label for="' + str(
    #                         personal.id) + str(personal.id) + '">' + str(personal.fullname) + '</label></div></li>'
    #                     html_e = html_e + html
    #
    #                     admin[em][su][ger] = html_e
    #     return admin

    def get_employees_tree(self):
        query = EmpresaManager(self.db).get_all()
        admin = dict()

        for empresa in query:
            em = (empresa.id, empresa.nombre)
            admin[em] = dict()

            list_ciudad = CiudadManager(self.db).get_all()
            cont_ciu = 1
            cont_suc = 1
            cont_ger = 1

            for ciudades in list_ciudad:
                ciu = (cont_ciu, ciudades.nombre)
                cont_ciu = cont_ciu + 1
                admin[em][ciu] = dict()

                list_sucursal = SucursalManager(self.db).listar_x_ciudad(ciudades.id)

                for sucursales in list_sucursal:

                    su = (cont_suc, sucursales.nombre)
                    cont_suc = cont_suc + 1
                    admin[em][ciu][su] = dict()

                    list_gerencias = SucursalManager(self.db).obtener_gerencias(empresa.id,sucursales.id)

                    for gerencias in list_gerencias:
                        ger = (cont_ger, list_gerencias[gerencias]['gerencia'])
                        cont_ger = cont_ger + 1
                        admin[em][ciu][su][ger] = dict()

                        list_personal = PersonaManager(self.db).obtener_x_gerencia(sucursales.id,list_gerencias[gerencias]['idgerencia'])
                        html_e = ""
                        for personal in list_personal:
                            html = '<li class="dd-item" data-id="' + str(personal.id)+ str(personal.id) + '"><div class="dd-handle"><input id="' + str(personal.id) + str(personal.id) + '" data-id="' + str(personal.id) + '" data-sex="' + str(personal.sexo) + '"type="checkbox" class="module chk-col-deep-purple employee"><label for="' + str(personal.id) + str(personal.id) + '">' + str(personal.fullname) + '</label></div></li>'
                            html_e = html_e + html

                            admin[em][ciu][su][ger] = html_e
        return admin

    def get_dataemp(self, idp):
        objeto = self.db.query(Persona).join(Empleado).filter(Persona.id == idp).order_by(Persona.fullname.asc()).all()

        return objeto

    def obtener_persona(self, id):
        objeto = self.db.query(Persona).filter(Persona.enabled == True).filter(Persona.id == id).first()

        return objeto

    def persona_excel(self, empleados):
        fecha = datetime.now()
        cname = "Empleados" + fecha.strftime('%Y-%m-%d') + ".xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte Empleados'

        indice = 0
        # --------------------------------------------------------------------
        indice = indice + 1
        ws['A' + str(indice)] = 'Nº'
        ws['B' + str(indice)] = 'PAIS'
        ws['C' + str(indice)] = 'DEPARTAMENTO'
        ws['D' + str(indice)] = 'CIUDAD'
        ws['E' + str(indice)] = 'EMPRESA'
        ws['F' + str(indice)] = 'SUCURSAL'
        ws['G' + str(indice)] = 'CENTRO DE COSTO'
        ws['H' + str(indice)] = 'SECCION'
        ws['I' + str(indice)] = 'CODIGO'
        ws['J' + str(indice)] = 'NOMBRE'
        ws['K' + str(indice)] = 'APELLIDO_P'
        ws['L' + str(indice)] = 'APELLIDO_M'
        ws['M' + str(indice)] = 'CI'
        ws['N' + str(indice)] = 'SEXO'
        ws['O' + str(indice)] = 'FECHA_NACIMIENTO'
        ws['P' + str(indice)] = 'CARGO'
        ws['Q' + str(indice)] = 'CONTRATO_PLAZO'
        ws['R' + str(indice)] = 'FECHA_INGRESO_NOM'
        ws['S' + str(indice)] = 'FECHA_VENCIMIENTO'
        ws['T' + str(indice)] = 'CORREO'

        ws['A' + str(indice)].font = Font(bold=True)
        ws['B' + str(indice)].font = Font(bold=True)
        ws['C' + str(indice)].font = Font(bold=True)
        ws['D' + str(indice)].font = Font(bold=True)
        ws['E' + str(indice)].font = Font(bold=True)
        ws['F' + str(indice)].font = Font(bold=True)
        ws['G' + str(indice)].font = Font(bold=True)
        ws['H' + str(indice)].font = Font(bold=True)
        ws['I' + str(indice)].font = Font(bold=True)
        ws['J' + str(indice)].font = Font(bold=True)
        ws['K' + str(indice)].font = Font(bold=True)
        ws['L' + str(indice)].font = Font(bold=True)
        ws['M' + str(indice)].font = Font(bold=True)
        ws['N' + str(indice)].font = Font(bold=True)
        ws['O' + str(indice)].font = Font(bold=True)
        ws['P' + str(indice)].font = Font(bold=True)
        ws['Q' + str(indice)].font = Font(bold=True)
        ws['R' + str(indice)].font = Font(bold=True)
        ws['S' + str(indice)].font = Font(bold=True)
        ws['T' + str(indice)].font = Font(bold=True)


        for i in PersonaManager(self.db).listar_todos():
            x = PersonaManager(self.db).get_dataemp(i.id)

            fechaf = ""
            if x[0].contrato[0].fechaFin:
                fechaf = x[0].contrato[0].fechaFin.strftime('%d/%m/%Y')


            indice = indice + 1
            ws['A' + str(indice)] = x[0].id
            ws['B' + str(indice)] = x[0].empleado[0].pais.nombre
            ws['C' + str(indice)] = x[0].empleado[0].departamentos.nombre
            ws['D' + str(indice)] = x[0].empleado[0].ciudad.nombre
            ws['E' + str(indice)] = x[0].empleado[0].sucursal.empresa.nombre
            ws['F' + str(indice)] = x[0].empleado[0].sucursal.nombre
            ws['G' + str(indice)] = x[0].empleado[0].centro.nombre
            ws['H' + str(indice)] = x[0].empleado[0].gerencia.nombre
            ws['I' + str(indice)] = x[0].empleado[0].codigo
            ws['J' + str(indice)] = x[0].nombres
            ws['K' + str(indice)] = x[0].apellidopaterno
            ws['L' + str(indice)] = x[0].apellidomaterno
            ws['M' + str(indice)] = x[0].ci
            ws['N' + str(indice)] = x[0].sexo
            ws['O' + str(indice)] = x[0].fechanacimiento.strftime('%d/%m/%Y')
            ws['P' + str(indice)] = x[0].empleado[0].cargo.nombre
            ws['Q' + str(indice)] = x[0].contrato[0].tipo
            ws['R' + str(indice)] = x[0].contrato[0].fechaIngreso.strftime('%d/%m/%Y')
            ws['S' + str(indice)] = fechaf
            ws['T' + str(indice)] = x[0].empleado[0].email


        wb.save("server/common/resources/downloads/persona/" + cname)
        return cname

    def get_all_data(self, id):
        list = {}
        c = 0
        for dataper in self.db.query(Persona).filter(Persona.enabled == True).filter(Persona.id == id):
            #list[c] = dict(id=dataper.ValorParametro.id, nombre=dataper.ValorParametro.nombre, valor=dataper.ValorParametro.valor)
            c = c + 1
        return list


class ContratoManager(SuperManager):
    def __init__(self, db):
        super().__init__(Contrato, db)

    def get_all(self):
        return self.db.query(self.entity).filter(self.entity.enabled == True).all()

    def list_all(self):
        return dict(objects=self.db.query(self.entity).join(Persona).filter(self.entity.enabled == True))

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.enabled == True)

    def obtener_concluido(self):
        return self.db.query(Contrato).join(Persona).filter(Contrato.enabled == False).all()

    def validar_estado(self):
        try:
            fecha_td = datetime.now(pytz.timezone('America/La_Paz'))
            fc_today = fecha_td.date()
            contratos = self.db.query(self.entity).filter(self.entity.enabled == True).filter(self.entity.tipo != 'INDEFINIDO').all()
            for con in contratos:
                fecha_fn = con.fechaFin.date()

                if fc_today > fecha_fn:
                    con.enabled = False
                    con.persona.enabled = False

            self.db.commit()
            return dict(message='success')
        except Exception as e:
            print(e)
            return dict(message='failed')


class AdministrativoManager(SuperManager):
    def __init__(self, db):
        super().__init__(Administrativo, db)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Administrativo.", fecha=fecha, tabla="rrhh_persona_administrativo", identificador=a.id)
        super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Administrativo.", fecha=fecha, tabla="rrhh_persona_administrativo", identificador=a.id)
        super().insert(b)
        return a


class EstudiosManager(SuperManager):
    def __init__(self, db):
        super().__init__(Estudios, db)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Estudio.", fecha=fecha, tabla="rrhh_persona_estudios", identificador=a.id)
        super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Estudio.", fecha=fecha, tabla="rrhh_persona_estudios", identificador=a.id)
        super().insert(b)
        return a


class CapacitacionManager(SuperManager):
    def __init__(self, db):
        super().__init__(Capacitacion, db)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Capacitación.", fecha=fecha, tabla="rrhh_persona_capacitacion", identificador=a.id)
        super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Capacitación.", fecha=fecha, tabla="rrhh_persona_capacitacion", identificador=a.id)
        super().insert(b)
        return a


class EducacionManager(SuperManager):
    def __init__(self, db):
        super().__init__(Educacion, db)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Educación.", fecha=fecha, tabla="rrhh_persona_educacion", identificador=a.id)
        super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Educación.", fecha=fecha, tabla="rrhh_persona_educacion", identificador=a.id)
        super().insert(b)
        return a


class MemoManager(SuperManager):
    def __init__(self, db):
        super().__init__(Memo, db)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Memo.", fecha=fecha, tabla="rrhh_persona_memo", identificador=a.id)
        super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Memo.", fecha=fecha, tabla="rrhh_persona_memo", identificador=a.id)
        super().insert(b)
        return a


class IdiomaManager(SuperManager):
    def __init__(self, db):
        super().__init__(Idioma, db)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Idioma.", fecha=fecha, tabla="rrhh_persona_idioma", identificador=a.id)
        super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Idioma.", fecha=fecha, tabla="rrhh_persona_idioma", identificador=a.id)
        super().insert(b)
        return a


class ExperienciaManager(SuperManager):
    def __init__(self, db):
        super().__init__(Experiencia, db)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Experiencia.", fecha=fecha, tabla="rrhh_persona_experiencia", identificador=a.id)
        super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Experiencia.", fecha=fecha, tabla="rrhh_persona_experiencia", identificador=a.id)
        super().insert(b)
        return a


class PadresManager(SuperManager):
    def __init__(self, db):
        super().__init__(Padres, db)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Padre.", fecha=fecha, tabla="rrhh_persona_padres", identificador=a.id)
        super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Padre.", fecha=fecha, tabla="rrhh_persona_padres", identificador=a.id)
        super().insert(b)
        return a


class ConyugeManager(SuperManager):
    def __init__(self, db):
        super().__init__(Conyuge, db)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Cónyuge.", fecha=fecha, tabla="rrhh_persona_conyuge", identificador=a.id)
        super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Cónyuge.", fecha=fecha, tabla="rrhh_persona_conyuge", identificador=a.id)
        super().insert(b)
        return a


class HijosManager(SuperManager):
    def __init__(self, db):
        super().__init__(Hijos, db)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Hijo.", fecha=fecha, tabla="rrhh_persona_hijos", identificador=a.id)
        super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Hijo.", fecha=fecha, tabla="rrhh_persona_hijos", identificador=a.id)
        super().insert(b)
        return a


class CoordenadasManager(SuperManager):
    def __init__(self, db):
        super().__init__(Coordenadas, db)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Coordenadas del personal.", fecha=fecha,
                     tabla="cb_rrhh_persona_coordenadas", identificador=a.id)
        super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Coordenadas del personal.", fecha=fecha,
                     tabla="cb_rrhh_persona_coordenadas", identificador=a.id)
        super().insert(b)
        return a
