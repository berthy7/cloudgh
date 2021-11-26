from ...operaciones.bitacora.managers import *
from server.common.managers import SuperManager
from .models import *
from sqlalchemy.exc import IntegrityError
from ...personal.persona.models import *
from ..solicitud.models import *
from ..historico.models import *

from sqlalchemy.sql import func,or_,and_
import decimal

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font



class V_personalManager(SuperManager):

    def __init__(self, db):
        super().__init__(V_personal, db)

    def obtener_x_fkpersona(self, idpersona):
        x = self.db.query(self.entity).filter(self.entity.fkpersona == idpersona).order_by(self.entity.gestion.desc()).all()

        return x

    def crear_reporte(self, per, diccionary):

        # fechainicio = datetime.strptime(diccionary['fechainicio'], '%d/%m/%Y')
        # fechafin = datetime.strptime(diccionary['fechafin'], '%d/%m/%Y')
        detalle = ""

        vacaciones_persona = V_personalManager(self.db).obtener_x_fkpersona(per)

        # lista_fechas = BitacoraManager(self.db).rango_fechas(fechainicio, fechafin)

        for vac in vacaciones_persona:


            fecha_gestion = datetime.strptime("01/01/"+str(vac.gestion), '%d/%m/%Y')
            fecha_gestion_final= datetime.strptime("31/12/" + str(vac.gestion), '%d/%m/%Y')


            dias = self.db.query(func.sum(V_solicitud.dias)).filter(V_solicitud.fkpersona == per).filter(and_(V_solicitud.estadoautorizacion == "Aceptado",V_solicitud.estadoaprobacion == "Aceptado")).filter(V_solicitud.fechai.between(fecha_gestion, fecha_gestion_final))

            if dias[0][0]:
                dias_vacaciones = dias[0][0]
            else:
                dias_vacaciones = 0


            dias_rechazo = self.db.query(func.sum(V_solicitud.dias)).filter(V_solicitud.fkpersona == per).filter(
                or_(V_solicitud.estadoautorizacion == "Rechazado", V_solicitud.estadoaprobacion == "Rechazado")).filter(
                V_solicitud.fechai.between(fecha_gestion, fecha_gestion_final))

            if dias_rechazo[0][0]:
                dias_vacaciones_rechazo = dias_rechazo[0][0]
            else:
                dias_vacaciones_rechazo = 0


            detalle = detalle + "" \
                                "<tr border='1'>" \
                                "<td colspan='2' align='left'><font size=3>" + str(vac.persona.fullname) + "</font></td>" \
                               "<td colspan='1' align='left'><font size=3>" + str(vac.gestion) + "</font></td>" \
                                "<td colspan='1' align='left'><font size=3>" + str(vac.dias) + " </font></td>" \
                               "<td colspan='1' align='left'><font size=3>" + str(dias_vacaciones) + " </font></td>" \
                             "<td colspan='1' align='left'><font size=3>" + str(dias_vacaciones_rechazo) + " </font></td>" \
                                   "</tr>"

        return detalle

    def get_all(self):
        return self.db.query(self.entity).filter(self.entity.enabled == True).all()

    def list_all(self):
        return dict(objects=self.db.query(self.entity).filter(self.entity.enabled == True))

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.enabled == True)

    def obtener_x_personal(self,fkpersona):
        dias = 0
        gestion = 0
        x = self.db.query(self.entity).filter(self.entity.fkpersona == fkpersona).first()

        if x:
            dias = x.dias
            gestion = x.gestion

        return dict(mensaje="", respuesta=True, tipo="", objeto=str(dias), gestion=str(gestion))

    def obtener_vacacion_disponible(self, fkpersona):
        dias = 0
        gestion = 0
        x = self.db.query(self.entity).filter(or_(self.entity.estado == "Disponible", self.entity.estado == "Diferida")).filter(self.entity.fkpersona == fkpersona).all()

        for i in x:
            dias = dias + i.dias

        return dict(dias=str(dias))

    def listar_x_ciudad(self, idciudad):
        return self.db.query(self.entity).filter(self.entity.fkciudad == idciudad).filter(
            self.entity.enabled == True)

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro V_personal.",
                     fecha=fecha, tabla="cb_vacaciones_personal", identificador=a.id)
        super().insert(b)
        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico V_personal.",
                     fecha=fecha, tabla="cb_vacaciones_personal", identificador=a.id)
        super().insert(b)
        return a

    def delete(self, id, estado, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.enabled = estado
        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion="Eliminó V_personal.", fecha=fecha,
                     tabla="rrhh_sucursal", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()
        return x

    def importar_excel(self, cname, user, ip):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['LEGAJO', 'NOMBRE','APELLIDO_P', 'APELLIDO_M','DIAS_VACACION','GESTION']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                for row in ws.iter_rows(min_row=2):
                    codigo = row[indices['LEGAJO']].value
                    dias_vacacion = row[indices['DIAS_VACACION']].value
                    gestion = row[indices['GESTION']].value


                    if codigo is not None and dias_vacacion is not None:

                        query_v_personal = self.db.query(V_personal).join(Persona).join(Empleado).filter(Empleado.codigo == str(codigo)).filter(V_personal.gestion == str(gestion)).first()

                        if not query_v_personal:

                            query = self.db.query(Persona).join(Empleado).filter(Empleado.codigo == str(codigo)).first()

                            if query:
                                fecha_zona = datetime.now(pytz.timezone('America/La_Paz'))
                                histori = V_historico(fkpersona=query.id, dias=dias_vacacion,descripcion="Importacion Excel gestion: " +str(gestion) ,operacion="+",fecha=fecha_zona)

                                self.db.add(histori)
                                self.db.flush()

                                fecha_hoy = datetime.now(pytz.timezone('America/La_Paz'))
                                fecha = fecha_hoy.year

                                resultado = int(fecha) - int(gestion)
                                estado = ""

                                if resultado == 0 or resultado == 1:
                                    estado = "Disponible"
                                if resultado >= 2:
                                    estado = "Diferida"


                                personal = V_personal(fkpersona=query.id, dias=dias_vacacion, gestion=gestion,estado = estado)
                                self.db.add(personal)
                                self.db.flush()
                    else:
                        self.db.rollback()
                        return {'message': 'Hay Columnas vacias', 'success': False}

                self.db.commit()

                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except IntegrityError as e:
            self.db.rollback()
            if 'UNIQUE constraint failed: rrhh_persona.dni' in str(e):
                return {'message': 'CI duplicado', 'success': False}
            if 'UNIQUE constraint failed: rrhh_empleado.codigo' in str(e):
                return {'message': 'codigo de empleado duplicado', 'success': False}
            return {'message': str(e), 'success': False}

    def actualizar_nro_vacacion(self, fkpersona, dias, operacion, user, ip,idSolicitud):
        fecha = datetime.now(pytz.timezone('America/La_Paz'))
        v_personal = self.db.query(V_personal).filter(V_personal.fkpersona == fkpersona).filter(V_personal.gestion == fecha.year).first()
        total_dias = 0
        lista_Gestion = list()

        if operacion == "+":
            total_dias = v_personal.dias + decimal.Decimal(dias)
            accion = "agrego dias de vacaciones"
            v_personal.dias = total_dias
            self.db.merge(v_personal)
            self.db.commit()

            b = Bitacora(fkusuario=user, ip=ip, accion=accion, fecha=fecha, tabla="cb_vacaciones_personal", identificador=v_personal.id)
            super().insert(b)
        elif operacion == "-":
            cant = decimal.Decimal(dias)

            for v in self.db.query(V_personal).filter(V_personal.fkpersona == fkpersona).order_by(V_personal.gestion.asc()).all():
                cantGestion = 0
                sw = False
                if cant == 0:
                    return lista_Gestion


                if v.dias > cant:
                    total_dias = v.dias - decimal.Decimal(cant)
                    cantGestion = cant
                    cant = 0
                    sw = True
                elif v.dias > 0:
                    total_dias = 0

                    cant = cant - v.dias
                    cantGestion = v.dias
                    sw = True

                if sw:
                    accion = "resto dias de vacaciones"
                    v.dias = total_dias
                    self.db.merge(v)
                    self.db.commit()

                    b = Bitacora(fkusuario=user, ip=ip, accion=accion, fecha=fecha, tabla="cb_vacaciones_personal", identificador=v_personal.id)
                    super().insert(b)

                    lista_Gestion.append(dict(fksolicitud=idSolicitud,dias=cantGestion,gestion=v.gestion,estado=v.estado))



            if cant > 0:
                total_dias = v_personal.dias - decimal.Decimal(cant)
                accion = "resto dias de vacaciones"
                v_personal.dias = total_dias
                self.db.merge(v_personal)
                self.db.commit()

                b = Bitacora(fkusuario=user, ip=ip, accion=accion, fecha=fecha, tabla="cb_vacaciones_personal", identificador=v_personal.id)
                super().insert(b)

                lista_Gestion.append(dict(fksolicitud=idSolicitud, dias=total_dias, gestion=v.gestion, estado=v.estado))

        return  lista_Gestion


    def disponibilidad(self, diccionario):
        fechai = datetime.strptime(diccionario['fechai'], '%d/%m/%Y')
        fechaf = datetime.strptime(diccionario['fechaf'], '%d/%m/%Y')
        lista_sin_vacacion = list()
        lista_sin_vacacion_autorizados = list()

        repetidos = set(diccionario['personas_arbol']).intersection(diccionario['personas'])

        for rep in repetidos:
            diccionario['personas'].remove(rep)

        for per in diccionario['personas']:
            diccionario['personas_arbol'].append(per)

        cant_dias = BitacoraManager(self.db).obtener_cant_dias(diccionario['fktipovacacion'], fechai, fechaf)

        if int(diccionario['fktipovacacion']) != 3:
            if int(diccionario['fktipovacacion']) == 4:
                return dict(mensaje="", respuesta=True, tipo="", lista_personas=lista_sin_vacacion)

            # n = self.db.query(V_personal).filter(V_personal.fkpersona == diccionario['fkpersona']).first()
            n = V_personalManager(self.db).obtener_vacacion_disponible(diccionario['fkpersona'])

            if n:
                if decimal.Decimal(n['dias']) != 0:
                    if cant_dias <= decimal.Decimal(n['dias']):
                        return dict(mensaje="", respuesta=True, tipo="", lista_personas=lista_sin_vacacion)
                    else:
                        return dict(mensaje="Tiene "+ str(n['dias']) +" dias", respuesta=False, tipo="warning", lista_personas="no")
                else:
                    return dict(mensaje="Tiene 0 dias", respuesta=False, tipo="warning", lista_personas=lista_sin_vacacion)
            else:
                return dict(mensaje="No tiene vacacion disponible", respuesta=False, tipo="error", lista_personas="no")
        else:
            for sin_vacacion in diccionario['sin_vacacion']:
                if sin_vacacion['estado']:
                    lista_sin_vacacion_autorizados.append(int(sin_vacacion['id']))
                else:
                    diccionario['personas_arbol'].remove(int(sin_vacacion['id']))

            for perso in diccionario['personas_arbol']:
                # n = self.db.query(V_personal).filter(V_personal.fkpersona == perso).first()
                n = V_personalManager(self.db).obtener_vacacion_disponible(perso)
                persona = self.db.query(Persona).filter(Persona.id == perso).first()
                print(n['dias'])

                if n:
                    if decimal.Decimal(n['dias']) != 0:
                        if cant_dias <= decimal.Decimal(n['dias']):
                            print("")
                            # return dict(mensaje="", respuesta=True, tipo="")
                        else:
                            if persona.id in lista_sin_vacacion_autorizados:
                                print("")
                            else:
                                lista_sin_vacacion.append(dict(id=persona.id,nombre=persona.fullname+" dias disponibles: "+n['dias']))
                    else:
                        if persona.id in lista_sin_vacacion_autorizados:
                            print("")
                        else:
                            lista_sin_vacacion.append(dict(id=persona.id, nombre=persona.fullname+" dias disponibles: "+n['dias']))
                else:
                    print("")
                    lista_sin_vacacion.append(persona.fullname)
                    # return dict(mensaje="No tiene vacacion disponible", respuesta=False, tipo="error")

            if len(lista_sin_vacacion) == 0:
                return dict(mensaje="", respuesta=True, tipo="", lista_personas=lista_sin_vacacion)
            else:
                return dict(mensaje="", respuesta=False, tipo="", lista_personas=lista_sin_vacacion)

    def dias(self, diccionario):

        fechafin = BitacoraManager(self.db).obtener_fechafin(diccionario)


        if 0 == 0:
            return dict(mensaje="", respuesta=True, tipo="",fechaf=fechafin)
        else:
            return dict(mensaje="", respuesta=False, tipo="",fechaf="")
